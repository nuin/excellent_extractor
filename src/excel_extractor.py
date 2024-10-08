import os
import logging
from dataclasses import dataclass
from typing import List, Tuple, Optional

import openpyxl
import xlrd
from openpyxl_image_loader import SheetImageLoader
import pytesseract
from PIL import Image
from whoosh import index
from whoosh.fields import Schema, TEXT, ID, STORED
from whoosh.qparser import QueryParser
from rich.console import Console
from rich.logging import RichHandler
from rich.progress import Progress
from rich.table import Table


@dataclass
class SheetContent:
    name: str
    cell_text: str
    images: List[Tuple[str, str]]  # (image_coord, image_text)


@dataclass
class WorkbookContent:
    filename: str
    relative_path: str
    sheets: List[SheetContent]


class ExcelExtractor:
    def __init__(self, directory_path: str, index_dir: str):
        self.directory_path = directory_path
        self.index_dir = index_dir
        self.schema = Schema(
            filename=ID(stored=True),
            relative_path=ID(stored=True),
            sheet_name=ID(stored=True),
            content=TEXT(stored=True),
            image_content=TEXT(stored=True)
        )
        self.console = Console()

    def extract_text_from_sheet(self, sheet, is_xlsx: bool) -> str:
        text = []
        if is_xlsx:
            for row in sheet.iter_rows(values_only=True):
                text.append('\t'.join(str(cell) if cell is not None else '' for cell in row))
        else:  # xls format
            for row in range(sheet.nrows):
                text.append('\t'.join(str(cell.value) if cell.value is not None else '' for cell in sheet.row(row)))
        return '\n'.join(text)

    def extract_text_from_image(self, image: Image) -> str:
        return pytesseract.image_to_string(image)

    def process_sheet(self, sheet, sheet_name: str, is_xlsx: bool) -> SheetContent:
        cell_text = self.extract_text_from_sheet(sheet, is_xlsx)

        images = []
        if is_xlsx:
            try:
                if not sheet.parent._read_only:  # Check if the workbook is not in read-only mode
                    image_loader = SheetImageLoader(sheet)
                    for image_coord in image_loader._images:
                        image = image_loader.get(image_coord)
                        image_text = self.extract_text_from_image(image)
                        images.append((image_coord, image_text))
            except Exception as e:
                logging.warning(f"Failed to extract images from sheet {sheet_name}. Error: {str(e)}")

        return SheetContent(sheet_name, cell_text, images)

    def process_workbook(self, file_path: str) -> Optional[WorkbookContent]:
        filename = os.path.basename(file_path)
        relative_path = os.path.relpath(file_path, self.directory_path)
        sheets = []

        try:
            if filename.endswith('.xlsx'):
                # First, try to load the workbook in read-only mode
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                is_xlsx = True
            elif filename.endswith('.xls'):
                wb = xlrd.open_workbook(file_path)
                is_xlsx = False
            else:
                raise ValueError("Unsupported file format")
        except Exception as e:
            if filename.endswith('.xlsx'):
                logging.warning(f"Failed to load {filename} in read-only mode. Trying normal mode. Error: {str(e)}")
                try:
                    # If read-only mode fails, try normal mode
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                except Exception as e:
                    logging.error(f"Failed to load {filename}. Error: {str(e)}")
                    return None
            else:
                logging.error(f"Failed to load {filename}. Error: {str(e)}")
                return None

        if is_xlsx:
            for sheet_name in wb.sheetnames:
                try:
                    sheet = wb[sheet_name]
                    sheet_content = self.process_sheet(sheet, sheet_name, is_xlsx)
                    sheets.append(sheet_content)
                except Exception as e:
                    logging.error(f"Failed to process sheet {sheet_name} in {filename}. Error: {str(e)}")
            wb.close()
        else:
            for sheet_index in range(wb.nsheets):
                try:
                    sheet = wb.sheet_by_index(sheet_index)
                    sheet_content = self.process_sheet(sheet, sheet.name, is_xlsx)
                    sheets.append(sheet_content)
                except Exception as e:
                    logging.error(f"Failed to process sheet {sheet.name} in {filename}. Error: {str(e)}")

        return WorkbookContent(filename, relative_path, sheets)

    def process_directory(self) -> List[WorkbookContent]:
        workbooks = []
        excel_files = []

        # Recursively search for Excel files
        for root, _, files in os.walk(self.directory_path):
            for file in files:
                if file.endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                    excel_files.append(os.path.join(root, file))

        with Progress() as progress:
            task = progress.add_task("[green]Processing Excel files...", total=len(excel_files))

            for file_path in excel_files:
                relative_path = os.path.relpath(file_path, self.directory_path)
                logging.info(f"Processing {relative_path}...")
                workbook_content = self.process_workbook(file_path)
                if workbook_content:
                    workbooks.append(workbook_content)
                progress.update(task, advance=1)

        return workbooks

    def index_content(self, workbooks: List[WorkbookContent]):
        if not os.path.exists(self.index_dir):
            os.mkdir(self.index_dir)
        ix = index.create_in(self.index_dir, self.schema)

        with Progress() as progress:
            task = progress.add_task("[green]Indexing content...", total=sum(len(wb.sheets) for wb in workbooks))

            writer = ix.writer()
            for workbook in workbooks:
                for sheet in workbook.sheets:
                    content = sheet.cell_text
                    image_content = "\n".join(image_text for _, image_text in sheet.images)
                    writer.add_document(
                        filename=workbook.filename,
                        relative_path=workbook.relative_path,
                        sheet_name=sheet.name,
                        content=content,
                        image_content=image_content
                    )
                    progress.update(task, advance=1)
            writer.commit()

    def search(self, query_str: str, limit: Optional[int] = None) -> List[dict]:
        ix = index.open_dir(self.index_dir)
        with ix.searcher() as searcher:
            query = QueryParser("content", ix.schema).parse(query_str)
            results = searcher.search(query, limit=limit)
            return [
                {
                    "filename": r['filename'],
                    "relative_path": r['relative_path'],
                    "sheet_name": r['sheet_name'],
                    "score": r.score,
                    "highlight": r.highlights("content")
                }
                for r in results
            ]

    def get_file_location(self, filename: str) -> Optional[dict]:
        ix = index.open_dir(self.index_dir)
        with ix.searcher() as searcher:
            query = QueryParser("filename", ix.schema).parse(filename)
            results = searcher.search(query, limit=1)
            if results:
                return {
                    'filename': results[0]['filename'],
                    'relative_path': results[0]['relative_path']
                }
        return None

    def search_images(self, query_str: str) -> List[dict]:
        ix = index.open_dir(self.index_dir)
        with ix.searcher() as searcher:
            query = QueryParser("image_content", ix.schema).parse(query_str)
            results = searcher.search(query)
            return [
                {
                    "filename": r['filename'],
                    "relative_path": r['relative_path'],
                    "sheet_name": r['sheet_name'],
                    "score": r.score,
                    "highlight": r.highlights("image_content")
                }
                for r in results
            ]

    def search_by_filename(self, filename: str) -> List[dict]:
        ix = index.open_dir(self.index_dir)
        with ix.searcher() as searcher:
            query = QueryParser("filename", ix.schema).parse(f"*{filename}*")
            results = searcher.search(query)
            return [
                {
                    'filename': r['filename'],
                    'relative_path': r['relative_path']
                }
                for r in results
            ]

    def search_by_gene_symbol(self, gene_symbol: str) -> List[dict]:
        gene_folder = os.path.join(self.directory_path, gene_symbol)
        if not os.path.exists(gene_folder):
            return []

        results = []
        for root, _, files in os.walk(gene_folder):
            for file in files:
                if file.endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                    relative_path = os.path.relpath(os.path.join(root, file), self.directory_path)
                    results.append({
                        'filename': file,
                        'relative_path': relative_path
                    })
        return results


def setup_logging(log_level: str = 'INFO'):
    logging.basicConfig(
        level=log_level,
        format="%(message)s",
        datefmt="[%X]",
        handlers=[RichHandler(rich_tracebacks=True)]
    )


if __name__ == "__main__":
    import click


    @click.group()
    @click.option('--log-level', type=click.Choice(['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL']), default='INFO',
                  help='Set the logging level')
    def cli(log_level):
        setup_logging(log_level)


    @cli.command()
    @click.option('--directory', type=click.Path(exists=True), required=True, help='Directory containing Excel files')
    @click.option('--index-dir', type=click.Path(), default='index_directory',
                  help='Directory to store the search index')
    def process(directory, index_dir):
        """Process Excel files and create search index"""
        extractor = ExcelExtractor(directory, index_dir)
        workbooks = extractor.process_directory()
        extractor.index_content(workbooks)
        click.echo(f"Processing and indexing completed. Index stored in: {index_dir}")


    @cli.command()
    @click.argument('query')
    @click.option('--index-dir', type=click.Path(exists=True), default='index_directory',
                  help='Directory where the search index is stored')
    @click.option('--limit', default=None, type=int,